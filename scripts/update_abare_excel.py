import sqlite3
import re
import openpyxl

DB_PATH = "agendha.db"
EXCEL_PATH = "abare_geral.xlsx"
OUTPUT_PATH = "app/abare_geral_atualizado.xlsx"

def normalize_cpf(cpf):
    if not cpf:
        return None
    # Remove everything that is NOT a digit
    return re.sub(r'\D', '', str(cpf))

def build_beneficiary_map(conn):
    """
    Fetches all beneficiaries and builds a map: {cleaned_cpf: (municipio, comunidade)}
    This ensures we match regardless of how CPF is stored in DB.
    """
    cursor = conn.cursor()
    mapping = {}
    
    # Try generic 'beneficiarios' table
    try:
        # Fetching id to ensure we have some data, but mainly cpf, mun, com
        cursor.execute("SELECT cpf, municipio, comunidade FROM beneficiarios")
        rows = cursor.fetchall()
        print(f"Loaded {len(rows)} records from 'beneficiarios' table.")
        
        for cpf, mun, com in rows:
            clean = normalize_cpf(cpf)
            if clean:
                mapping[clean] = (mun, com)
                
    except sqlite3.OperationalError as e:
        print(f"Warning: Could not read 'beneficiarios': {e}")
        
    # Try 'agua_que_alimenta_beneficiarios' if it exists (previous error suggested it might not, but let's be safe)
    try:
        cursor.execute("SELECT cpf, municipio, comunidade FROM agua_que_alimenta_beneficiarios")
        rows = cursor.fetchall()
        print(f"Loaded {len(rows)} records from 'agua_que_alimenta_beneficiarios' table.")
        
        for cpf, mun, com in rows:
            clean = normalize_cpf(cpf)
            if clean:
                # Update/Overwrite with more specific table if needed
                mapping[clean] = (mun, com)
    except sqlite3.OperationalError:
        pass
        
    return mapping

def main():
    print(f"Connecting to database: {DB_PATH}")
    conn = sqlite3.connect(DB_PATH)
    
    print("Building in-memory CPF map from database...")
    db_map = build_beneficiary_map(conn)
    print(f"Total unique CPFs in DB map: {len(db_map)}")
    
    if not db_map:
        print("Error: No data found in database to map.")
        return

    print(f"Loading Excel: {EXCEL_PATH}")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    except FileNotFoundError:
        print(f"Error: File {EXCEL_PATH} not found.")
        return

    # Find headers
    header_row = 1
    cpf_col = None
    mun_col = None
    com_col = None
    
    max_col = ws.max_column
    
    # Identify columns
    for col in range(1, max_col + 1):
        cell_value = ws.cell(row=header_row, column=col).value
        if not cell_value:
            continue
        v = str(cell_value).lower().strip()
        if 'cpf' in v:
            cpf_col = col
        elif 'município' in v or 'municipio' in v:
            mun_col = col
        elif 'comunidade' in v:
            com_col = col

    if not cpf_col:
        print("Error: CPF column not found in Excel.")
        return

    # Create columns if not exist
    if not mun_col:
        mun_col = max_col + 1
        ws.cell(row=header_row, column=mun_col, value="Município")
        max_col += 1
        print("Created column 'Município'")
        
    if not com_col:
        com_col = max_col + 1
        ws.cell(row=header_row, column=com_col, value="Comunidade")
        max_col += 1
        print("Created column 'Comunidade'")

    print(f"Excel Columns - CPF: {cpf_col}, Município: {mun_col}, Comunidade: {com_col}")

    updated_count = 0
    not_found_count = 0
    
    # Iterate all rows
    print("Processing rows...")
    for row in range(header_row + 1, ws.max_row + 1):
        cpf_cell = ws.cell(row=row, column=cpf_col)
        raw_cpf = cpf_cell.value
        
        # Clean Excel CPF
        clean_cpf = normalize_cpf(raw_cpf)
        
        if not clean_cpf:
            continue
            
        # Lookup in DB map
        if clean_cpf in db_map:
            municipio, comunidade = db_map[clean_cpf]
            
            # Update cells
            # Only update if value is present in DB, otherwise leave blank or existing?
            # User said: "Preencher Município e Comunidade... independente de como estão escritos"
            # It implies overwriting or filling.
            
            if municipio:
                ws.cell(row=row, column=mun_col, value=municipio)
            if comunidade:
                ws.cell(row=row, column=com_col, value=comunidade)
                
            updated_count += 1
        else:
            not_found_count += 1
            # Optional: print first few missing checks
            if not_found_count <= 5:
                print(f"CPF mismatch: {clean_cpf} (Original: {raw_cpf})")

    wb.save(OUTPUT_PATH)
    print(f"Saved updated file to: {OUTPUT_PATH}")  # nosec
    print(f"Total rows updated: {updated_count}")  # nosec
    print(f"CPFs not found in DB: {not_found_count}")

if __name__ == "__main__":
    main()
