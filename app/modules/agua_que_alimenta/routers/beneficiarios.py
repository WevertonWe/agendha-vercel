import logging
import datetime
import io
import asyncio
from typing import List, Union, Optional
from collections import defaultdict
import pandas as pd
from jinja2 import Environment, FileSystemLoader  # noqa: E402

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Response, Request
from fastapi.responses import JSONResponse, HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel

from app.core.database import get_supabase, fetch_all
from app.core.auth.dependencies import get_admin_user
from app.modules.agua_que_alimenta.models import (
    BeneficiarioUpdate, BeneficiarioParaKML, SchemaValidacao
)
from app.config import settings
from app.services.utils import remover_acentos, limpar_cpf
from app.services import ai_vision

router = APIRouter(prefix="/api/beneficiarios", tags=["Beneficiários"])
router_root = APIRouter() # Roteador auxiliar para rotas raiz (sem prefixo acumulado)

# --- Models para Relatórios ---
class RelatorioRequest(BaseModel):
    ids: List[int]
    colunas: List[str]
    email: Optional[str] = None

_env = Environment(loader=FileSystemLoader("app/templates"), cache_size=0)
templates = Jinja2Templates(env=_env)
logger = logging.getLogger(__name__)

@router_root.get("/beneficiarios/perfil/{id}", response_class=HTMLResponse)
async def ver_perfil_beneficiario(request: Request, id: int):
    """
    Exibe a página de perfil detalhado de um beneficiário específico via Supabase.
    """
    try:
        supabase = get_supabase()
        res = supabase.table('beneficiarios').select('*').eq('id', id).execute()
        
        if not res.data:
            return templates.TemplateResponse("errors/404.html", {"request": request}, status_code=404)
            
        beneficiario = res.data[0]
        
        return templates.TemplateResponse("beneficiarios/perfil.html", {
            "request": request,
            "beneficiario": beneficiario
        })
    except Exception as e:
        logger.error(f"Erro ao carregar perfil: {e}")
        return templates.TemplateResponse("errors/500.html", {"request": request, "detail": str(e)}, status_code=500)

@router_root.get("/planejamento/abare", response_class=HTMLResponse)
async def view_planejamento_abare(request: Request):
    return templates.TemplateResponse("agua/planejamento_abare.html", {"request": request})

@router.get("", response_class=JSONResponse)
def get_beneficiarios(
    municipio: str | None = None
):
    """
    Retorna a lista JSON de todos os beneficiários usando Supabase com paginação segura.
    """
    try:
        registros = fetch_all('beneficiarios')
        
        if municipio:
            registros = [r for r in registros if str(r.get('municipio') or '').strip().upper() == municipio.upper()]
            
        # Sanitização preventiva de campos de data
        for r in registros:
            for key in r.keys():
                if 'data' in key or 'updated' in key or 'created' in key:
                    r[key] = str(r[key]) if r[key] is not None else ''
                    
        return registros
    except Exception as e:
        logging.error(f"API: Erro ao buscar beneficiários no Supabase: {e}")
        raise HTTPException(status_code=500, detail="Erro interno ao buscar dados.")



@router.post("/import/confirmar")
async def confirmar_importacao_csv(
    dados: Union[List[dict], dict]
):
    try:
        supabase = get_supabase()
        count_novos = 0
        count_atualizados = 0

        lista_dados = dados if isinstance(dados, list) else [dados]

        for item in lista_dados:
            cpf_bruto = item.get('cpf')
            cpf_limpo = limpar_cpf(cpf_bruto)
            
            if not cpf_limpo:
                continue
            
            nome = item.get('nome', 'Desconhecido')
            c_comunidade = item.get('comunidade')
            c_municipio = remover_acentos(item.get('municipio'))
            c_nis = item.get('nis')
            c_lat = item.get('latitude')
            c_lon = item.get('longitude')
            c_status = 'IMPORTADO'
            
            # Busca por CPF para decidir entre insert ou update (Conforme pedido, edições devem usar ID se possível,
            # mas na importação o CPF é a chave de busca única)
            res = supabase.table('beneficiarios').select('id').eq('cpf', cpf_limpo).execute()
            
            payload = {
                "nome_completo": nome,
                "nome_familiar": nome,
                "comunidade": c_comunidade,
                "municipio": c_municipio,
                "nis": c_nis,
                "latitude": c_lat,
                "longitude": c_lon,
                "status": c_status
            }
            
            if res.data:
                # Usa o ID encontrado para atualizar
                beneficiario_id = res.data[0]['id']
                supabase.table('beneficiarios').update(payload).eq('id', beneficiario_id).execute()
                count_atualizados += 1
            else:
                # Insere novo
                payload.update({
                    "cpf": cpf_limpo,
                    "cpf_familiar": cpf_limpo,
                    "doc_status": 'PENDENTE'
                })
                supabase.table('beneficiarios').insert(payload).execute()
                count_novos += 1

        return {"message": "Processamento concluído", "novos": count_novos, "atualizados": count_atualizados}

    except Exception as e:
        logging.error(f"Erro ao confirmar importação CSV: {e}", exc_info=True)
        return JSONResponse(content={"error": f"Erro interno ao processar importação: {str(e)}"}, status_code=500)


@router.delete("/{beneficiario_id}", status_code=204)
def excluir_beneficiario(beneficiario_id: int):
    try:
        supabase = get_supabase()
        res = supabase.table('beneficiarios').delete().eq('id', beneficiario_id).execute()
        if not res.data:
            return JSONResponse(content={"error": "Beneficiário não encontrado."}, status_code=404)
        return Response(status_code=204)
    except Exception as e:
        logging.error(f"Erro ao excluir ID {beneficiario_id}: {e}")
        return JSONResponse(content={"error": f"Erro ao excluir: {str(e)}"}, status_code=500)


@router.get("/{beneficiario_id}", response_class=JSONResponse)
def get_beneficiario_por_id(beneficiario_id: int):
    try:
        supabase = get_supabase()
        res = supabase.table('beneficiarios').select('*').eq('id', beneficiario_id).execute()
        if not res.data:
            return JSONResponse(content={"error": "Beneficiário não encontrado."}, status_code=404)
        return res.data[0]
    except Exception as e:
        logging.error(f"Erro ao buscar ID {beneficiario_id}: {e}")
        return JSONResponse(content={"error": f"Erro no servidor: {str(e)}"}, status_code=500)


@router.put("/{beneficiario_id}", response_class=JSONResponse)
def update_beneficiario(beneficiario_id: int, dados: BeneficiarioUpdate):
    try:
        campos_para_atualizar = dados.dict(exclude_unset=True)
        if not campos_para_atualizar:
            return JSONResponse(content={"error": "Nenhum dado fornecido para atualização."}, status_code=400)

        # Sanitização de CPF se estiver presente
        if 'cpf' in campos_para_atualizar:
            cpf_limpo = limpar_cpf(campos_para_atualizar['cpf'])
            if cpf_limpo:
                campos_para_atualizar['cpf'] = cpf_limpo
                campos_para_atualizar['cpf_familiar'] = cpf_limpo
            else:
                return JSONResponse(content={"error": "CPF inválido."}, status_code=400)

        supabase = get_supabase()
        res = supabase.table('beneficiarios').update(campos_para_atualizar).eq('id', beneficiario_id).execute()
        
        if not res.data:
            return JSONResponse(content={"error": "Beneficiário não encontrado."}, status_code=404)
            
        return res.data[0]

    except Exception as e:
        logging.error(f"Erro ao atualizar ID {beneficiario_id}: {e}")
        return JSONResponse(content={"error": f"Erro inesperado no servidor: {str(e)}"}, status_code=500)


@router.post("/{beneficiario_id}/documento")
async def upload_documento_beneficiario(
    beneficiario_id: int,
    arquivo: UploadFile = File(...)
):
    try:
        supabase = get_supabase()
        # Nomenclatura fixa doc_{id}.pdf para permitir overwrite via upsert
        ext = arquivo.filename.split('.')[-1] if '.' in arquivo.filename else 'pdf'
        nome_arquivo = f"doc_{beneficiario_id}.{ext}"
        content = await arquivo.read()

        supabase.storage.from_('agendha-uploads').upload(
            path=f"beneficiarios_docs/{nome_arquivo}",
            file=content,
            file_options={"content-type": arquivo.content_type or "application/pdf", "upsert": "true"}
        )

        public_url = supabase.storage.from_('agendha-uploads').get_public_url(f"beneficiarios_docs/{nome_arquivo}")

        res = supabase.table('beneficiarios').update({"doc_status": public_url}).eq('id', beneficiario_id).execute()
        
        if not res.data:
            return JSONResponse(content={"error": "Beneficiário não encontrado."}, status_code=404)
            
        return res.data[0]

    except Exception as e:
        logging.error(f"Erro no upload de documento ID {beneficiario_id}: {e}")
        return JSONResponse(content={"error": f"Erro ao fazer upload do documento: {str(e)}"}, status_code=500)


@router.patch("/{beneficiario_id}/desvincular-pedreiro", status_code=200)
def desvincular_pedreiro(beneficiario_id: int):
    try:
        supabase = get_supabase()
        res = supabase.table('beneficiarios').update({"pedreiro_id": None}).eq('id', beneficiario_id).execute()
        if not res.data:
            return JSONResponse(content={"error": "Beneficiário não encontrado."}, status_code=404)
            
        return {"message": "Pedreiro desvinculado com sucesso."}
    except Exception as e:
        logging.error(f"Erro ao desvincular pedreiro ID {beneficiario_id}: {e}")
        return JSONResponse(content={"error": f"Erro ao desvincular pedreiro: {str(e)}"}, status_code=500)


@router.post("/{beneficiario_id}/nota_fiscal")
async def upload_nota_fiscal(
    beneficiario_id: int,
    arquivo: UploadFile = File(...)
):
    try:
        supabase = get_supabase()
        ext = arquivo.filename.split('.')[-1] if '.' in arquivo.filename else 'pdf'
        nome_arquivo = f"nf_{beneficiario_id}.{ext}"
        content = await arquivo.read()

        supabase.storage.from_('agendha-uploads').upload(
            path=f"pedreiros_docs/{nome_arquivo}",
            file=content,
            file_options={"content-type": arquivo.content_type or "application/pdf", "upsert": "true"}
        )

        public_url = supabase.storage.from_('agendha-uploads').get_public_url(f"pedreiros_docs/{nome_arquivo}")

        res = supabase.table('beneficiarios').update({
            "link_nota_fiscal": public_url,
            "status_pagamento": 'PAGO'
        }).eq('id', beneficiario_id).execute()
        
        if not res.data:
            return JSONResponse(content={"error": "Beneficiário não encontrado."}, status_code=404)
            
        return res.data[0]

    except Exception as e:
        logging.error(f"Erro no upload de nota fiscal ID {beneficiario_id}: {e}")
        return JSONResponse(content={"error": f"Erro ao fazer upload da nota fiscal: {str(e)}"}, status_code=500)


@router.post("/export/kml", response_class=Response)
async def exportar_beneficiarios_kml_post(beneficiarios_filtrados: List[BeneficiarioParaKML]):
    """
    Exportação KML via POST (Lista Explícita).
    
    Recebe uma lista JSON de beneficiários (filtrados no frontend) e gera o KML.
    Útil quando o filtro é complexo ou selecionado manualmente pelo usuário.
    """
    if not beneficiarios_filtrados:
        raise HTTPException(
            status_code=400, detail="Nenhum beneficiário fornecido para exportação.")

    status_para_estilo = {
        'EM CADASTRO': 'style_em_cadastro',
        'CADASTRADO': 'style_cadastrado',
        'A CONSTRUIR': 'style_a_construir',
        'CONSTRUÍDA': 'style_construida',
        'DEFAULT': 'style_default'
    }

    kml_styles = """
        <Style id="style_em_cadastro">
            <IconStyle>
                <Icon><href>http://maps.google.com/mapfiles/kml/paddle/ylw-stars.png</href></Icon>
                <scale>1.0</scale>
            </IconStyle>
            <LabelStyle><scale>0.8</scale></LabelStyle>
        </Style>
        <Style id="style_cadastrado">
            <IconStyle>
                <Icon><href>http://maps.google.com/mapfiles/kml/paddle/blu-circle.png</href></Icon>
                 <scale>1.0</scale>
            </IconStyle>
             <LabelStyle><scale>0.8</scale></LabelStyle>
        </Style>
        <Style id="style_a_construir">
            <IconStyle>
                <Icon><href>http://maps.google.com/mapfiles/kml/paddle/wht-blank.png</href></Icon>
                 <scale>1.0</scale>
            </IconStyle>
             <LabelStyle><scale>0.8</scale></LabelStyle>
        </Style>
         <Style id="style_construida">
            <IconStyle>
                <Icon><href>http://maps.google.com/mapfiles/kml/paddle/grn-stars.png</href></Icon>
                 <scale>1.0</scale>
            </IconStyle>
             <LabelStyle><scale>0.8</scale></LabelStyle>
        </Style>
        <Style id="style_default">
            <IconStyle>
                <Icon><href>http://maps.google.com/mapfiles/kml/paddle/red-circle.png</href></Icon>
                 <scale>1.0</scale>
            </IconStyle>
             <LabelStyle><scale>0.8</scale></LabelStyle>
        </Style>
    """

    placemarks = []
    for ben in beneficiarios_filtrados:
        if not ben.latitude or not ben.longitude:
            continue

        try:
            lat = float(str(ben.latitude).replace(',', '.'))
            lon = float(str(ben.longitude).replace(',', '.'))

            # Relaxed Validation: Accept any valid coordinate (Brazil roughly -34 to +5 N, -74 to -34 W)
            # Warning if EXACTLY 0.0 (often default value)
            if lat == 0.0 and lon == 0.0:
                 logging.warning(f"Beneficiário {ben.id} ignorado: Coordenadas Zero.")
                 continue

        except (ValueError, TypeError):
            # logging.debug(f"Beneficiário {ben.id} ignorado: Coordenadas Inválidas ({ben.latitude}, {ben.longitude})")
            continue

        status_normalizado = str(
            ben.status).upper() if ben.status else 'DEFAULT'
        style_id = status_para_estilo.get(
            status_normalizado, status_para_estilo['DEFAULT'])

        description = f"""
            <![CDATA[
                <b>Nome:</b> {ben.nome_completo or 'Não informado'}<br/>
                <b>CPF:</b> {ben.cpf or 'Não informado'}<br/>
                <b>Comunidade:</b> {ben.comunidade or 'Não informada'}<br/>
                <b>Status:</b> {ben.status or 'Não informado'}
            ]]>
        """

        placemark = f"""
            <Placemark>
                <name>{ben.nome_completo or 'Beneficiário sem nome'}</name>
                <description>{description}</description>
                <styleUrl>#{style_id}</styleUrl>
                <Point>
                    <coordinates>{lon},{lat},0</coordinates> 
                </Point>
            </Placemark>
        """
        placemarks.append(placemark)

    if not placemarks:
        raise HTTPException(
            status_code=400, detail="Nenhum beneficiário com coordenadas válidas encontrado para exportar.")

    kml_content = f"""<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2">
    <Document>
        <name>Beneficiarios Agua que Alimenta</name>
        {kml_styles}
        {"".join(placemarks)}
    </Document>
</kml>
    """

    filename = f"beneficiarios_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.kml"  # noqa: F841

    return Response(
        content=kml_content,
        media_type='application/vnd.google-earth.kml+xml',
    )


@router.post("/import/comparar", response_class=JSONResponse)
async def comparar_importacao_csv(
    file: UploadFile = File(...),
    current_user = Depends(get_admin_user)
):
    """
    Recebe um CSV/Excel, localiza colunas dinamicamente e retorna lista consolidada para triagem via Supabase.
    """
    import csv

    content = await file.read()
    filename = file.filename.lower()
    
    try:
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            df = pd.read_excel(io.BytesIO(content))
        else:
            decoded = content.decode('utf-8-sig')
            csv_reader = csv.DictReader(io.StringIO(decoded), delimiter=';')
            if len(csv_reader.fieldnames or []) < 2:
                csv_reader = csv.DictReader(io.StringIO(decoded), delimiter=',')
            df = pd.DataFrame(list(csv_reader))

        # Robust Column Search Helper
        def find_column(df, candidates):
            cols_map = {str(c).strip().upper(): c for c in df.columns}
            for candidate in candidates:
                cand_upper = candidate.upper()
                if cand_upper in cols_map:
                    return cols_map[cand_upper]
                for real_col_upper, real_col_name in cols_map.items():
                    if cand_upper in real_col_upper:
                         return real_col_name
            return None

        col_cpf = find_column(df, ['CPF', 'C.P.F', 'DOC', 'DOCUMENTO', 'NR_CPF'])
        col_nome = find_column(df, ['NOME', 'NOME_COMPLETO', 'BENEFICIARIO', 'CANDIDATO'])
        col_nis = find_column(df, ['NIS', 'PIS', 'NIT'])
        col_comunidade = find_column(df, ['COMUNIDADE', 'LOCALIDADE', 'POVOADO'])
        col_lat = find_column(df, ['LAT', 'LATITUDE', 'COORD_X'])
        col_lon = find_column(df, ['LON', 'LONG', 'LONGITUDE', 'COORD_Y'])
        col_municipio = find_column(df, ['MUNICIPIO', 'CIDADE', 'MUNICÍPIO'])
        col_status = find_column(df, ['STATUS', 'SITUACAO', 'ESTADO'])

        # Pre-fetch CPFs do Supabase
        supabase_data = fetch_all('beneficiarios', 'cpf, nome_completo, status, nis, comunidade, latitude, longitude')
        db_beneficiarios = {r['cpf']: r for r in supabase_data if r.get('cpf')}

        lista_triagem = []

        for index, row in df.iterrows():
            def get_val(col_name):
                if not col_name: return ""  # noqa: E701
                val = row[col_name]
                return str(val).strip() if pd.notna(val) else ""

            raw_nome = get_val(col_nome) or "Desconhecido"
            raw_cpf = get_val(col_cpf)
            raw_nis = get_val(col_nis)
            raw_com = get_val(col_comunidade)
            raw_lat = get_val(col_lat)
            raw_lon = get_val(col_lon)
            raw_status = get_val(col_status).upper()
            raw_mun = remover_acentos(get_val(col_municipio))

            cpf_limpo = limpar_cpf(raw_cpf)

            item = {
                "id_temp": f"temp_{index}",
                "nome": raw_nome,
                "cpf": cpf_limpo or "", 
                "cpf_original": raw_cpf,
                "nis": raw_nis,
                "comunidade": raw_com,
                "municipio": raw_mun,
                "latitude": raw_lat,
                "longitude": raw_lon,
                "status": raw_status,
                "status_triagem": "INVALIDO",
                "msg_erro": "",
                "dados_banco": None,
                "diffs": []
            }

            if not cpf_limpo or len(cpf_limpo) != 11:
                item["status_triagem"] = "INVALIDO"
                item["msg_erro"] = "CPF Inválido ou Vazio"
            elif cpf_limpo in db_beneficiarios:
                existing = db_beneficiarios[cpf_limpo]
                item["dados_banco"] = existing
                item["status_triagem"] = "DUPLICADO"
                
                diffs = []
                def check_diff(field_csv, field_db, label):
                    val_csv = str(item[field_csv] or "").strip().upper()
                    val_db = str(existing.get(field_db) or "").strip().upper()
                    if val_csv and val_csv != val_db:
                        diffs.append({"campo": label, "novo": item[field_csv], "antigo": existing.get(field_db)})

                check_diff("nome", "nome_completo", "Nome")
                check_diff("nis", "nis", "NIS")
                check_diff("comunidade", "comunidade", "Comunidade")
                
                val_status_csv = raw_status.strip().upper()
                val_status_db = str(existing.get('status') or "").strip().upper()
                if val_status_csv and val_status_csv != val_status_db:
                    diffs.append({"campo": "Status", "novo": raw_status, "antigo": existing.get('status')})

                if diffs:
                    item["diffs"] = diffs
            else:
                item["status_triagem"] = "NOVO"
                if not raw_nome or len(raw_nome) < 3 or raw_nome == "Desconhecido":
                     item["status_triagem"] = "ATENCAO" 
                     item["msg_erro"] = "Nome ausente ou muito curto"

            lista_triagem.append(item)

        return {
            "resumo": {
                "total": len(lista_triagem),
                "novos": sum(1 for i in lista_triagem if i['status_triagem'] == 'NOVO'),
                "duplicados": sum(1 for i in lista_triagem if i['status_triagem'] == 'DUPLICADO'),
                "invalidos": sum(1 for i in lista_triagem if i['status_triagem'] in ['INVALIDO', 'ATENCAO'])
            },
            "lista_triagem": lista_triagem
        }

    except Exception as e:
        logging.error(f"Erro ao comparar importação CSV: {e}")
        return JSONResponse(content={"error": f"Erro ao processar arquivo: {str(e)}"}, status_code=400)


@router.get("/consolidado/atividades", response_class=JSONResponse)
def get_consolidado_atividades():
    try:
        supabase = get_supabase()
        res = supabase.table('beneficiarios').select('municipio, status').execute()
        
        if not res.data:
            return []
            
        registros = res.data
        dados_agregados = defaultdict(lambda: defaultdict(int))
        for registro in registros:
            mun = str(registro.get("municipio") or '').strip().upper()
            if not mun:
                continue
                
            status = str(registro.get("status") or '').strip().upper()
            if status == 'EXCLUIDO':
                continue
                
            dados_agregados[mun]["total_beneficiarios"] += 1
            if status in ['EM CADASTRO', 'EM_CADASTRO']:
                dados_agregados[mun]["em_cadastro"] += 1
            elif status in ['CADASTRADO']:
                dados_agregados[mun]["cadastrado"] += 1
            elif status in ['A CONSTRUIR', 'A_CONSTRUIR']:
                dados_agregados[mun]["a_construir"] += 1
            elif status in ['CONSTRUÍDA', 'CONSTRUIDA', 'CONCLUÍDO', 'CONCLUIDO']:
                dados_agregados[mun]["construida"] += 1
            else:
                dados_agregados[mun]["outros_status"] += 1

        dados_consolidados = []
        status_keys = ["em_cadastro", "cadastrado", "a_construir", "construida", "outros_status"]
        
        for nome, contadores in sorted(dados_agregados.items()):
            resultado_final = {
                'municipio': nome,
                'total_beneficiarios': int(contadores.get('total_beneficiarios', 0))
            }
            for key in status_keys:
                resultado_final[key] = int(contadores.get(key, 0))

            dados_consolidados.append(resultado_final)

        return dados_consolidados

    except Exception as e:
        logging.error(f"API: Erro ao gerar dados consolidados no Supabase: {e}")
        raise HTTPException(status_code=500, detail="Erro interno ao consolidar dados.")


@router.get("/municipios", response_class=JSONResponse)
def get_municipios_unicos():
    try:
        supabase = get_supabase()
        res = supabase.table('beneficiarios').select('municipio').execute()
        if res.data:
            municipios = {row.get('municipio') for row in res.data if row.get('municipio')}
            return sorted(list(municipios))
        return []
    except Exception as e:
        logging.error(f"Erro ao buscar municípios no Supabase: {e}")
        return []

@router.get("/comunidades", response_class=JSONResponse)
def get_comunidades_unicas():
    try:
        supabase = get_supabase()
        res = supabase.table('beneficiarios').select('comunidade').execute()
        if res.data:
            comunidades = {row.get('comunidade') for row in res.data if row.get('comunidade')}
            return sorted(list(comunidades))
        return []
    except Exception as e:
        logging.error(f"Erro ao buscar comunidades no Supabase: {e}")
        return []


@router.post("/fix/sync-cpf", status_code=200)
def fix_sync_cpf_familiar(current_user = Depends(get_admin_user)):
    """
    Script de Correção: Iguala cpf_familiar ao cpf via Supabase.
    """
    try:
        supabase = get_supabase()
        # Busca registros onde CPF Familiar está diferente do CPF
        res = supabase.table('beneficiarios').select('id, cpf').execute()
        
        count = 0
        for ben in res.data:
            if ben.get('cpf'):
                supabase.table('beneficiarios').update({"cpf_familiar": ben['cpf']}).eq('id', ben['id']).execute()
                count += 1
                
        return {"message": f"Sucesso! {count} beneficiários corrigidos no Supabase."}
    except Exception as e:
        logging.error(f"Erro ao sincronizar CPF Familiar: {e}")
        return JSONResponse(content={"error": str(e)}, status_code=500)


@router_root.post("/api/salvar-validado", status_code=200)
def salvar_validacao(dados: SchemaValidacao):
    import os
    from pathlib import Path
    
    cpf_limpo = limpar_cpf(dados.cpf)
    if not cpf_limpo:
        return JSONResponse(content={"error": "CPF inválido ou não informado."}, status_code=400)

    id_fila_para_baixa = dados.id_fila or dados.id
        
    try:
        supabase = get_supabase()
        
        if id_fila_para_baixa:
            try:
                from app.services import store
                
                item_fila = store.get_item(str(id_fila_para_baixa))
                if item_fila and item_fila.get('caminho_arquivo_local'):
                    caminho_orig = item_fila['caminho_arquivo_local'] 
                    nome_orig = os.path.basename(caminho_orig)
                    
                    orig_path = Path(settings.UPLOAD_FOLDER) / nome_orig
                    if not orig_path.exists():
                         import tempfile
                         orig_path = Path(tempfile.gettempdir()) / nome_orig
                    
                    if orig_path.exists():
                        with open(orig_path, "rb") as f:
                            file_bytes = f.read()
                            
                        # Nomenclatura fixa doc_{id}.pdf será feita APÓS o insert/get_id se for novo
                        # Mas aqui temos um problema: precisamos do ID do Supabase.
                        # Vamos fazer o insert primeiro e depois o upload.
            except Exception as e_pdf:
                logging.error(f"Erro ao preparar PDF do OCR: {e_pdf}")

        dados_salvar = {
            'cpf': cpf_limpo,
            'cpf_familiar': cpf_limpo,
            'nome_completo': dados.nome_completo,
            'nome_familiar': dados.nome_completo,
            'data_nascimento': dados.data_nascimento,
            'escolaridade': dados.escolaridade,
            'comunidade': dados.comunidade,
            'municipio': dados.municipio,
            'nis': dados.nis,
            'estado_uf': dados.uf or dados.estado_uf,
            'ref_localizacao': dados.ref_localizacao,
            'sexo': dados.sexo,
            'status': "CADASTRADO"
        }
        
        dados_salvar = {k: v for k, v in dados_salvar.items() if v is not None}

        # 1. UPSERT no Supabase
        res_existente = supabase.table('beneficiarios').select('id').eq('cpf', cpf_limpo).execute()
        
        if res_existente.data:
            id_ben = res_existente.data[0]['id']
            supabase.table('beneficiarios').update(dados_salvar).eq('id', id_ben).execute()
        else:
            res_ins = supabase.table('beneficiarios').insert(dados_salvar).execute()
            id_ben = res_ins.data[0]['id']

        # 2. Upload do documento agora que temos o ID definitivo
        if id_fila_para_baixa and 'file_bytes' in locals():
            nome_fixo = f"doc_{id_ben}.pdf"
            supabase.storage.from_('agendha-uploads').upload(
                path=f"beneficiarios_docs/{nome_fixo}",
                file=file_bytes,
                file_options={"content-type": "application/pdf", "upsert": "true"}
            )
            doc_status_final = supabase.storage.from_('agendha-uploads').get_public_url(f"beneficiarios_docs/{nome_fixo}")
            supabase.table('beneficiarios').update({"doc_status": doc_status_final}).eq('id', id_ben).execute()

        if id_fila_para_baixa:
            from app.services import store
            store.delete_item(str(id_fila_para_baixa))
            
        return {"message": "Salvo com sucesso!", "id": id_ben}

    except Exception as e:
        logging.error(f"Erro ao salvar validação OCR: {e}")
        return JSONResponse(content={"error": f"Erro ao salvar: {str(e)}"}, status_code=500)
@router.get("/export/kml")
def exportar_beneficiarios_kml(
    municipio: str | None = None,
    status: str | None = None,
    comunidade: str | None = None
):
    """
    Exportação KML via Supabase.
    """
    try:
        supabase = get_supabase()
        query = supabase.table('beneficiarios').select('nome_completo, cpf, comunidade, latitude, longitude, status, municipio')
        
        if municipio:
            query = query.ilike('municipio', municipio)
        if status:
            query = query.ilike('status', status)
        if comunidade:
            query = query.ilike('comunidade', f"%{comunidade}%")
            
        res = query.execute()
        rows = res.data
        
        # Build KML
        kml = ['<?xml version="1.0" encoding="UTF-8"?>']
        kml.append('<kml xmlns="http://www.opengis.net/kml/2.2">')
        kml.append('<Document>')
        kml.append(f'<name>Beneficiarios_{municipio or "Geral"}</name>')
        
        # Define Style
        kml.append('<Style id="beneficiario">')
        kml.append('<IconStyle>')
        kml.append('<scale>1.2</scale>')
        kml.append('<Icon><href>http://maps.google.com/mapfiles/kml/shapes/man.png</href></Icon>')
        kml.append('</IconStyle>')
        kml.append('</Style>')
        
        for row in rows:
            lat = row.get('latitude')
            lon = row.get('longitude')
            if not lat or not lon or str(lat) == '0' or str(lon) == '0':
                continue

            nome = row.get('nome_completo') or "Sem Nome"
            cpf = row.get('cpf') or "N/I"
            com = row.get('comunidade') or "N/I"
            mun = row.get('municipio') or "N/I"
            
            description = f"<b>Comunidade:</b> {com}<br/><b>CPF:</b> {cpf}<br/><b>Município:</b> {mun}"
            
            kml.append('<Placemark>')
            kml.append(f'<name>{nome}</name>')
            kml.append(f'<description><![CDATA[{description}]]></description>')
            kml.append('<styleUrl>#beneficiario</styleUrl>')
            kml.append('<Point>')
            kml.append(f'<coordinates>{lon},{lat},0</coordinates>')
            kml.append('</Point>')
            kml.append('</Placemark>')
            
        kml.append('</Document>')
        kml.append('</kml>')
        
        kml_content = "\n".join(kml)
        filename = f"Export_Beneficiarios_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.kml"
        
        return Response(
            content=kml_content,
            media_type="application/vnd.google-earth.kml+xml",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logging.error(f"Erro ao exportar KML: {e}")
        return JSONResponse(content={"error": f"Erro na exportação: {str(e)}"}, status_code=500)

# ==============================================================================
# NOVAS ROTAS: GERADOR DE RELATÓRIOS CUSTOMIZADOS (PROJETO AGENDHA)
# ==============================================================================

@router_root.post("/api/relatorios/excel")
async def gerar_relatorio_excel(data: RelatorioRequest):
    """
    Gera um arquivo Excel via Supabase com colunas customizadas.
    """
    if not data.ids:
        return JSONResponse(content={"error": "Nenhum ID fornecido"}, status_code=400)
    
    try:
        supabase = get_supabase()
        
        colunas_permitidas = {
            "nome_familiar": "Nome",
            "cpf_familiar": "CPF",
            "municipio": "Município",
            "comunidade": "Comunidade",
            "status": "Status",
            "nis": "NIS",
            "tecnico_agua_que_alimenta": "Técnico",
            "doc_status": "Documento",
            "grh": "GRH"
        }
        
        cols_to_query = [c for c in data.colunas if c in colunas_permitidas]
        if not cols_to_query:
            cols_to_query = list(colunas_permitidas.keys())

        # Busca no Supabase
        res = supabase.table('beneficiarios').select(','.join(cols_to_query)).in_('id', data.ids).execute()
        dados_brutos = res.data
        
        dados = []
        for i, linha in enumerate(dados_brutos, start=1):
            if 'doc_status' in linha:
                val = linha['doc_status']
                linha['doc_status'] = 'OK' if (val and isinstance(val, str) and ('/' in val or '.pdf' in val.lower())) else 'Procurar documento'
            
            if 'numero_ordem' in data.colunas:
                nova_linha = {'Nº': i}
                nova_linha.update(linha)
                dados.append(nova_linha)
            else:
                dados.append(linha)

        df = pd.DataFrame(dados)
        df.rename(columns={k: v for k, v in colunas_permitidas.items() if k in cols_to_query}, inplace=True)

        output = io.BytesIO()
        from openpyxl.styles import PatternFill, Font
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Beneficiários')
            worksheet = writer.sheets['Beneficiários']
            
            for cell in worksheet[1]:
                if cell.value: cell.font = Font(bold=True) # noqa: E701
            
            for col in worksheet.columns:
                max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                worksheet.column_dimensions[col[0].column_letter].width = max_length + 2

            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for r in worksheet.iter_rows():
                for cell in r:
                    if cell.value == 'Procurar documento':
                        cell.fill = yellow_fill

        output.seek(0)
        return StreamingResponse(output, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': 'attachment; filename="relatorio_beneficiarios.xlsx"'})

    except Exception as e:
        logging.error(f"Erro ao gerar Excel: {e}")
        return JSONResponse(content={"error": f"Erro interno no Excel: {str(e)}"}, status_code=500)


@router_root.post("/api/relatorios/analise")
async def gerar_analise_ia(data: RelatorioRequest):
    """
    Análise de IA via Gemini usando dados do Supabase.
    """
    if not data.ids:
        return JSONResponse(content={"error": "Nenhum ID fornecido"}, status_code=400)

    try:
        supabase = get_supabase()
        cols_for_ai = ["nome_familiar", "municipio", "comunidade", "status", "nis", "tecnico_agua_que_alimenta"]
        res = supabase.table('beneficiarios').select(','.join(cols_for_ai)).in_('id', data.ids).execute()
        
        if not res.data:
            return {"analise": "Nenhum dado encontrado para os IDs informados."}

        df = pd.DataFrame(res.data)
        dados_json = df.to_json(orient="records", force_ascii=False)
        
        prompt = f"""
        Analise a seguinte lista de beneficiários ({len(df)} registros) do projeto Agendha:
        1. Resumo por Município/Comunidade.
        2. Distribuição de Status.
        3. Gargalos ou dados faltantes.
        4. Sugestão de foco imediato.
        Responda em Markdown elegante com emojis.
        DADOS: {dados_json}
        """

        client = ai_vision.get_gemini_client()
        if client:
            response = await asyncio.to_thread(client.models.generate_content, model="gemini-2.0-flash", contents=[prompt])
            return {"analise": response.text}
        else:
            return JSONResponse(content={"error": "API Gemini não configurada."}, status_code=500)

    except Exception as e:
        logging.error(f"Erro na análise IA: {e}")
        return JSONResponse(content={"error": f"Erro interno na IA: {str(e)}"}, status_code=500)
