"""
BSF Beneficiários - API Router
Endpoints para listagem, filtros e preparação de arquivos do módulo BSF.
"""
import logging
from typing import Optional

from fastapi import APIRouter, HTTPException, Query, File, UploadFile, Form
from fastapi.responses import JSONResponse, FileResponse
from pydantic import BaseModel
import io
import csv
import os
import unicodedata

from app.core.database import get_supabase
from app.services.utils import limpar_cpf

class BeneficiarioBSFCreate(BaseModel):
    nome_completo: str
    cpf: Optional[str] = None
    caf: Optional[str] = None
    municipio: Optional[str] = None
    comunidade: Optional[str] = None
    tecnico: Optional[str] = None

class AtividadeCreate(BaseModel):
    tipo_atividade: str
    data: str
    iniciativas_vinculadas: Optional[list] = None

class AtividadeUpdate(BaseModel):
    tipo_atividade: str
    data: str
    iniciativas_vinculadas: Optional[list] = None

class TipoAtividadeCreate(BaseModel):
    nome: str

class TipoAtividadeUpdate(BaseModel):
    nome: Optional[str] = None
    ativo: Optional[bool] = None

router = APIRouter(prefix="/api/bsf/beneficiarios", tags=["BSF Beneficiários"])

logger = logging.getLogger(__name__)


@router.get("/tipos-atividade")
async def listar_tipos_atividade(somente_ativos: bool = True):
    """Retorna a lista de tipos de atividade (ativos ou todos)."""
    try:
        supabase = get_supabase()
        query = supabase.table("bsf_tipos_atividade").select("*")
        if somente_ativos:
            query = query.eq("ativo", True)
        res = query.order("nome", desc=False).execute()
        return res.data or []
    except Exception as e:
        logger.error(f"Erro ao listar tipos de atividade BSF: {e}")
        raise HTTPException(status_code=500, detail="Erro ao carregar categorias de atividades.")

@router.post("/tipos-atividade")
async def criar_tipo_atividade(dados: TipoAtividadeCreate):
    """Cria um novo tipo de atividade."""
    try:
        supabase = get_supabase()
        nome_limpo = dados.nome.strip()
        if not nome_limpo:
            raise HTTPException(status_code=400, detail="O nome do tipo de atividade não pode ser vazio.")
        
        # Verificar se já existe (mesmo inativo)
        res_check = supabase.table("bsf_tipos_atividade").select("*").ilike("nome", nome_limpo).execute()
        if res_check.data:
            existente = res_check.data[0]
            # Se estiver inativo, reativar
            if not existente.get("ativo"):
                res_upd = supabase.table("bsf_tipos_atividade").update({"ativo": True, "nome": nome_limpo}).eq("id", existente["id"]).execute()
                return res_upd.data[0]
            raise HTTPException(status_code=400, detail="Esta categoria já existe e está ativa.")
            
        payload = {"nome": nome_limpo, "ativo": True}
        res = supabase.table("bsf_tipos_atividade").insert(payload).execute()
        return res.data[0]
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Erro ao criar tipo de atividade: {e}")
        raise HTTPException(status_code=500, detail="Erro ao criar tipo de atividade.")

@router.put("/tipos-atividade/{id}")
async def atualizar_tipo_atividade(id: int, dados: TipoAtividadeUpdate):
    """Atualiza o nome e/ou status ativo de um tipo de atividade."""
    try:
        supabase = get_supabase()
        payload = {}
        if dados.nome is not None:
            nome_limpo = dados.nome.strip()
            if not nome_limpo:
                raise HTTPException(status_code=400, detail="O nome não pode ser vazio.")
            payload["nome"] = nome_limpo
        if dados.ativo is not None:
            payload["ativo"] = dados.ativo
            
        if not payload:
            raise HTTPException(status_code=400, detail="Nenhum dado fornecido para atualização.")
            
        res = supabase.table("bsf_tipos_atividade").update(payload).eq("id", id).execute()
        if not res.data:
            raise HTTPException(status_code=404, detail="Tipo de atividade não encontrado.")
        return res.data[0]
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Erro ao atualizar tipo de atividade {id}: {e}")
        raise HTTPException(status_code=500, detail="Erro ao atualizar tipo de atividade.")

@router.delete("/tipos-atividade/{id}")
async def deletar_tipo_atividade(id: int):
    """Soft delete: desativa a categoria de atividade no banco."""
    try:
        supabase = get_supabase()
        res = supabase.table("bsf_tipos_atividade").update({"ativo": False}).eq("id", id).execute()
        if not res.data:
            raise HTTPException(status_code=404, detail="Tipo de atividade não encontrado.")
        return {"status": "success", "message": "Categoria desativada com sucesso."}
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Erro ao desativar tipo de atividade {id}: {e}")
        raise HTTPException(status_code=500, detail="Erro ao desativar tipo de atividade.")


def parse_links_safe(paths, supabase, context_info: str = "") -> list:
    """
    Parses and sanitizes document links safely to prevent AttributeError or list parsing errors.
    If paths is a JSON string, tries to load it. Discards non-string/null elements, logs a warning
    for invalid links, and generates storage public URLs safely using try-except.
    """
    if not paths:
        return []
        
    # Coerce to list if paths is a string
    if isinstance(paths, str):
        paths_trimmed = paths.strip()
        if paths_trimmed.startswith("[") and paths_trimmed.endswith("]"):
            import json
            try:
                paths = json.loads(paths_trimmed)
            except Exception as e:
                logger.warning(f"Falha ao decodificar array JSON de links '{paths_trimmed}' {context_info}: {e}")
                paths = [paths_trimmed]
        else:
            paths = [paths_trimmed]
            
    if not isinstance(paths, list):
        logger.warning(f"Formato de link inválido (não é lista nem string) {context_info}: {type(paths)}")
        return []
        
    sanitized_links = []
    for p in paths:
        if p is None:
            continue
        if not isinstance(p, str):
            logger.warning(f"Elemento de link ignorado por não ser string {context_info}: {p} (tipo {type(p)})")
            continue
            
        p_str = p.strip()
        if not p_str:
            continue
            
        if p_str.startswith("http"):
            sanitized_links.append(p_str)
        else:
            # Assumir que é um caminho do storage. Usar try-except preventivo com signed url.
            try:
                # Gera uma URL assinada válida por 1 ano (31536000 segundos)
                res_url = supabase.storage.from_("agendha-uploads").create_signed_url(p_str, 31536000)
                url_final = res_url if isinstance(res_url, str) else res_url.get("signedUrl")
                sanitized_links.append(url_final)
            except Exception as ex:
                logger.warning(f"Erro ao gerar URL assinada para '{p_str}', usando fallback público {context_info}: {ex}")
                # Fallback de segurança para URL pública caso o método falhe
                try:
                    public_url = supabase.storage.from_("agendha-uploads").get_public_url(p_str)
                    if public_url:
                        sanitized_links.append(public_url)
                except Exception as ex_fallback:
                    logger.warning(f"Erro no fallback público de '{p_str}' {context_info}: {ex_fallback}")
                
    return sanitized_links


@router.get("")
async def listar_beneficiarios(
    tecnico: Optional[str] = Query(None, description="Filtro por técnico responsável"),
    municipio: Optional[str] = Query(None, description="Filtro por município"),
    status: Optional[str] = Query(None, description="Filtro por status"),
    page: int = Query(1, ge=1, description="Página atual"),
    page_size: int = Query(50, ge=1, le=200, description="Itens por página"),
):
    """Lista beneficiários com filtros opcionais por técnico e município."""
    try:
        supabase = get_supabase()
        
        cols = (
            "id, nome_completo, cpf, caf, nis, municipio, comunidade, "
            "nome_tecnico, tecnico_agua_que_alimenta, status, "
            "verificado_bsf, data_atividade, projeto"
        )
        
        query = supabase.table("beneficiarios").select(cols, count="exact").eq("projeto", "Bahia Sem Fome")
        
        if tecnico:
            query = query.ilike("nome_tecnico", f"%{tecnico}%")
        if municipio:
            query = query.ilike("municipio", f"%{municipio}%")
        if status:
            query = query.ilike("status", f"%{status}%")
        
        # BSF filter: only verified beneficiaries if needed
        # query = query.not_.is_("verificado_bsf", "null")
        
        offset = (page - 1) * page_size
        query = query.range(offset, offset + page_size - 1)
        query = query.order("id", desc=True)
        
        res = query.execute()
        
        total = res.count if res.count is not None else len(res.data or [])
        
        return {
            "data": res.data or [],
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": max(1, -(-total // page_size)),
        }
        
    except Exception as e:
        logger.error(f"Erro ao listar beneficiários BSF: {e}")
        raise HTTPException(status_code=500, detail="Erro ao buscar beneficiários.")


@router.get("/filtros")
async def obter_filtros():
    """Retorna valores únicos de técnico e município para popular os selects de filtro."""
    try:
        supabase = get_supabase()
        
        # Fetch unique tecnicos
        res_tec = supabase.table("beneficiarios").select("nome_tecnico").eq("projeto", "Bahia Sem Fome").not_.is_("nome_tecnico", "null").execute()
        tecnicos_raw = [r["nome_tecnico"] for r in (res_tec.data or []) if r.get("nome_tecnico")]
        tecnicos = sorted(set(t.strip() for t in tecnicos_raw if t.strip()))
        
        # Fetch unique municipios
        res_mun = supabase.table("beneficiarios").select("municipio").eq("projeto", "Bahia Sem Fome").not_.is_("municipio", "null").execute()
        municipios_raw = [r["municipio"] for r in (res_mun.data or []) if r.get("municipio")]
        municipios = sorted(set(m.strip() for m in municipios_raw if m.strip()))
        
        # Fetch unique statuses
        res_st = supabase.table("beneficiarios").select("status").eq("projeto", "Bahia Sem Fome").not_.is_("status", "null").execute()
        statuses_raw = [r["status"] for r in (res_st.data or []) if r.get("status")]
        statuses = sorted(set(s.strip() for s in statuses_raw if s.strip()))
        
        return {
            "tecnicos": tecnicos,
            "municipios": municipios,
            "statuses": statuses,
        }
        
    except Exception as e:
        logger.error(f"Erro ao obter filtros BSF: {e}")
        raise HTTPException(status_code=500, detail="Erro ao buscar filtros.")


@router.get("/{beneficiario_id}")
async def detalhar_beneficiario(beneficiario_id: int):
    """Retorna os dados completos de um beneficiário específico."""
    try:
        supabase = get_supabase()
        res = supabase.table("beneficiarios").select("*").eq("id", beneficiario_id).eq("projeto", "Bahia Sem Fome").execute()
        
        if not res.data:
            raise HTTPException(status_code=404, detail="Beneficiário não encontrado.")
        
        return res.data[0]
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao detalhar beneficiário {beneficiario_id}: {e}")
        raise HTTPException(status_code=500, detail="Erro ao buscar beneficiário.")


@router.get("/{beneficiario_id}/atividades")
async def listar_atividades_beneficiario(beneficiario_id: int):
    """Retorna as atividades associadas a um beneficiário."""
    try:
        supabase = get_supabase()
        res = supabase.table("bsf_atividades").select("*").eq("beneficiario_id", beneficiario_id).order("data_atividade", desc=True).execute()
        
        # Parse arrays if they are returned as None
        atividades = res.data or []
        for a in atividades:
            context = f"da atividade {a.get('id')} do beneficiário {beneficiario_id}"
            a["link_sigater"] = parse_links_safe(a.get("link_sigater"), supabase, f"sigater {context}")
            a["link_colletum"] = parse_links_safe(a.get("link_colletum"), supabase, f"colletum {context}")
            a["link_ateste"] = parse_links_safe(a.get("link_ateste"), supabase, f"ateste {context}")
            
            # Retrocompatibilidade no response para o frontend (que lê atv.data e iniciativas_vinculadas)
            a["data"] = a.get("data_atividade")
            a["iniciativas_vinculadas"] = a.get("iniciativas_vinculadas") or []
            
        return atividades
    except Exception as e:
        logger.error(f"Erro ao listar atividades {beneficiario_id}: {e}")
        raise HTTPException(status_code=500, detail="Erro ao buscar atividades.")

@router.get("/{beneficiario_id}/metas")
async def listar_metas_beneficiario(beneficiario_id: int):
    """Retorna as metas e iniciativas associadas a um beneficiário com tratamento Null-Safety."""
    try:
        supabase = get_supabase()
        res = supabase.table("bsf_metas_plano").select("*").eq("beneficiario_id", beneficiario_id).execute()
        metas = res.data or []
        
        # Mapeamento preventivo contra nulos para evitar Erro 500 no template (Null-Safety)
        safe_metas = []
        for m in metas:
            safe_metas.append({
                "id": m.get("id") or 0,
                "beneficiario_id": m.get("beneficiario_id") or 0,
                "codigo": m.get("codigo") or "",
                "tipo": m.get("tipo") or "",
                "descricao": m.get("descricao") or "",
                "tarefa": m.get("tarefa") or "",
                "como_fazer": m.get("como_fazer") or ""
            })
        
        # Ordenação natural pelo código
        import re
        def natural_key(code):
            return [int(c) if c.isdigit() else c for c in re.split(r'(\d+)', code or "")]
            
        safe_metas.sort(key=lambda x: natural_key(x.get("codigo", "")))
        return safe_metas
    except Exception as e:
        logger.error(f"Erro ao listar metas {beneficiario_id}: {e}")
        raise HTTPException(status_code=500, detail="Erro ao buscar metas do plano produtivo.")

@router.post("/{beneficiario_id}/atividades")
async def criar_atividade(beneficiario_id: int, dados: AtividadeCreate):
    """Cria uma nova atividade para o beneficiário."""
    try:
        supabase = get_supabase()
        payload = {
            "beneficiario_id": beneficiario_id,
            "tipo_atividade": dados.tipo_atividade,
            "data_atividade": dados.data,
            "iniciativas_vinculadas": dados.iniciativas_vinculadas or [],
            "link_sigater": [],
            "link_colletum": [],
            "link_ateste": []
        }
        res = supabase.table("bsf_atividades").insert(payload).execute()
        return res.data[0]
    except Exception as e:
        logger.error(f"Erro ao criar atividade: {e}")
        raise HTTPException(status_code=500, detail="Erro ao criar atividade.")

@router.put("/atividades/{atividade_id}")
async def atualizar_atividade(atividade_id: int, dados: AtividadeUpdate):
    """Atualiza o tipo, a data e as iniciativas vinculadas da atividade específica."""
    try:
        supabase = get_supabase()
        payload = {
            "tipo_atividade": dados.tipo_atividade,
            "data_atividade": dados.data,
            "iniciativas_vinculadas": dados.iniciativas_vinculadas or []
        }
        res = supabase.table("bsf_atividades").update(payload).eq("id", atividade_id).execute()
        if not res.data:
            raise HTTPException(status_code=404, detail="Atividade não encontrada.")
        return res.data[0]
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Erro ao atualizar atividade {atividade_id}: {e}")
        raise HTTPException(status_code=500, detail="Erro ao atualizar atividade.")

@router.delete("/atividades/{atividade_id}")
async def deletar_atividade(atividade_id: int):
    """Remove a atividade específica do banco."""
    try:
        supabase = get_supabase()
        res = supabase.table("bsf_atividades").delete().eq("id", atividade_id).execute()
        if not res.data:
            raise HTTPException(status_code=404, detail="Atividade não encontrada.")
        return {"status": "success", "message": "Atividade removida com sucesso."}
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Erro ao deletar atividade {atividade_id}: {e}")
        raise HTTPException(status_code=500, detail="Erro ao remover atividade.")

@router.post("/{beneficiario_id}/atividades/{atividade_id}/upload")
async def upload_documento_atividade(
    beneficiario_id: int,
    atividade_id: int,
    tipo: str = Form(...),
    file: UploadFile = File(...)
):
    """Realiza o upload de um documento para a atividade específica."""
    try:
        if tipo not in ["sigater", "colletum", "ateste"]:
            raise HTTPException(status_code=400, detail="Tipo de documento inválido.")
            
        supabase = get_supabase()
        res_ben = supabase.table("beneficiarios").select("cpf, caf, municipio").eq("id", beneficiario_id).execute()
        if not res_ben.data:
            raise HTTPException(status_code=404, detail="Beneficiário não encontrado.")
            
        ben = res_ben.data[0]
        cpf_ou_caf = (ben.get("cpf") or ben.get("caf") or "SEM_DOC").replace(".", "").replace("-", "").replace(" ", "")
        municipio = (ben.get("municipio") or "SEM_MUNICIPIO").strip()
        
        nome_arquivo = f"{file.filename}"
        caminho_storage = f"bsf/{municipio}/{cpf_ou_caf}/atividade_{atividade_id}/{tipo}/{nome_arquivo}"
        
        file_content = await file.read()
        
        supabase.storage.from_("agendha-uploads").upload(
            file=file_content,
            path=caminho_storage,
            file_options={"content-type": "application/pdf"}
        )
        
        # Atualizar a tabela de atividades
        coluna = f"link_{tipo}"
        res_atv = supabase.table("bsf_atividades").select(coluna).eq("id", atividade_id).execute()
        
        if res_atv.data:
            links_atuais = res_atv.data[0].get(coluna) or []
            if caminho_storage not in links_atuais:
                links_atuais.append(caminho_storage)
                res_upd = supabase.table("bsf_atividades").update({coluna: links_atuais}).eq("id", atividade_id).execute()
                
                # Fetch signed URLs for the response so frontend can use them immediately
                updated_data = res_upd.data[0] if res_upd.data else None
                if updated_data:
                    for t in ["sigater", "colletum", "ateste"]:
                        context = f"da atividade {atividade_id} no upload do beneficiário {beneficiario_id}"
                        updated_data[f"link_{t}"] = parse_links_safe(updated_data.get(f"link_{t}"), supabase, f"{t} {context}")
                    
                return {"status": "success", "caminho": caminho_storage, "atividade": updated_data}
                
        return {"status": "success", "caminho": caminho_storage}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro no upload BSF: {e}")
        raise HTTPException(status_code=500, detail=f"Erro no upload: {e}")

@router.post("")
async def criar_beneficiario(dados: BeneficiarioBSFCreate):
    """Cria um novo beneficiário individualmente no projeto BSF."""
    try:
        supabase = get_supabase()
        cpf_limpo = limpar_cpf(dados.cpf) if dados.cpf else None
        caf_limpo = dados.caf.strip() if dados.caf else None
        
        if not cpf_limpo and not caf_limpo:
            raise HTTPException(400, "CPF ou CAF é obrigatório.")
            
        payload = {
            "nome_completo": dados.nome_completo,
            "cpf": cpf_limpo,
            "caf": caf_limpo,
            "municipio": dados.municipio,
            "comunidade": dados.comunidade,
            "nome_tecnico": dados.tecnico,
            "projeto": "Bahia Sem Fome",
            "status": "CADASTRADO"
        }
        res = supabase.table("beneficiarios").insert(payload).execute()
        return {"message": "Beneficiário cadastrado", "data": res.data[0]}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao cadastrar BSF: {e}")
        raise HTTPException(500, f"Erro ao cadastrar: {e}")

@router.delete("/{id}")
async def deletar_beneficiario(id: int):
    """Exclui permanentemente um beneficiário do projeto BSF."""
    print(f"DEBUG: Tentativa de exclusão do ID {id}")
    try:
        supabase = get_supabase()
        
        # Verificar se o beneficiário existe e pertence ao BSF
        res = supabase.table("beneficiarios").select("id").eq("id", id).eq("projeto", "Bahia Sem Fome").execute()
        if not res.data:
            raise HTTPException(status_code=404, detail="Beneficiário não encontrado.")
            
        # Deletar no Supabase (ON DELETE CASCADE no banco removerá os registros em bsf_atividades)
        supabase.table("beneficiarios").delete().eq("id", id).execute()
        
        return {"message": "Beneficiário excluído com sucesso."}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erro ao excluir beneficiário BSF: {e}")
        raise HTTPException(status_code=500, detail="Erro ao excluir beneficiário.")

@router.get("/template")
async def baixar_template():
    """Baixa a planilha modelo para importação."""
    caminho = os.path.join("app", "modules", "bahia_sem_fome", "assets", "planilha_exemplo.xlsx")
    if not os.path.exists(caminho):
        raise HTTPException(status_code=404, detail="Template não encontrado.")
    return FileResponse(
        path=caminho,
        filename="planilha_exemplo_bsf.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@router.post("/importar")
async def importar_planilha_bsf(file: UploadFile = File(...)):
    """Importa lista de beneficiários BSF em massa, detectando colunas."""
    try:
        content = await file.read()
        filename = file.filename.lower()
        
        linhas = []
        fieldnames = []
        
        if filename.endswith('.xlsx'):
            import openpyxl
            from zipfile import BadZipFile
            try:
                wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
                sheet = wb.active
            except BadZipFile as bz:
                logger.error(f"❌ [INVESTIGAÇÃO] Arquivo XLSX corrompido fisicamente: {bz}")
                print(f"❌ [INVESTIGAÇÃO] XLSX CORROMPIDO ZIP: {bz}")
                return JSONResponse(content={"error": f"Arquivo Excel corrompido (Zip inválido): {str(bz)}"}, status_code=400)
            except Exception as ex:
                logger.error(f"❌ [INVESTIGAÇÃO] Erro desconhecido no parser XLSX: {ex}")
                print(f"❌ [INVESTIGAÇÃO] PARSER XLSX FALHOU: {ex}")
                return JSONResponse(content={"error": f"Erro ao ler Excel: {str(ex)}"}, status_code=400)
            
            # Pegando cabeçalhos na primeira linha
            headers = [cell.value for cell in sheet[1]]
            fieldnames = [str(h).strip() for h in headers if h is not None]
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Ignora linhas totalmente vazias
                if any(row):
                    row_dict = {}
                    for i, header in enumerate(fieldnames):
                        if i < len(row):
                            # Preserva números como string para não quebrar validações
                            val = row[i]
                            row_dict[header] = str(val).strip() if val is not None else None
                    linhas.append(row_dict)
                    
        elif filename.endswith('.xls'):
            return JSONResponse(content={"error": "Formato .xls legado não suportado. Use .xlsx ou .csv."}, status_code=400)
            
        else:
            try:
                decoded = content.decode('utf-8-sig')
            except UnicodeDecodeError as ude:
                logger.error(f"❌ [INVESTIGAÇÃO] Erro de decodificação UTF-8 no CSV: {ude}")
                print(f"❌ [INVESTIGAÇÃO] UTF-8 DECODE FAILED: {ude}")
                # Tenta decodificar em ISO-8859-1 como fallback de segurança
                try:
                    decoded = content.decode('iso-8859-1')
                    logger.warning("🔍 [INVESTIGAÇÃO] Fallback para decodificação ISO-8859-1 obteve sucesso.")
                except Exception:
                    return JSONResponse(content={"error": f"Erro de codificação de caracteres no CSV: {str(ude)}"}, status_code=400)
                    
            try:
                sniffer = csv.Sniffer()
                try:
                    dialect = sniffer.sniff(decoded[:1024], delimiters=[',', ';', '\t'])
                    delimiter = dialect.delimiter
                except csv.Error as ce:
                    logger.warning(f"⚠️ [INVESTIGAÇÃO] Sniffer falhou em detectar delimitador, usando delimitador de segurança: {ce}")
                    delimiter = ';' if ';' in decoded[:100] else ','
                
                csv_reader = csv.DictReader(io.StringIO(decoded), delimiter=delimiter)
                linhas = list(csv_reader)
                fieldnames = csv_reader.fieldnames or []
            except csv.Error as csv_err:
                logger.error(f"❌ [INVESTIGAÇÃO] Falha estrutural no processador CSV: {csv_err}")
                print(f"❌ [INVESTIGAÇÃO] PARSER CSV FALHOU: {csv_err}")
                return JSONResponse(content={"error": f"Erro na estrutura interna do arquivo CSV: {str(csv_err)}"}, status_code=400)

        def normalizar_nome(t):
            if not t:
                return ""
            t_sem_acentos = remove_accents(t)
            return " ".join(t_sem_acentos.upper().split())

        def padronizar_cpf(raw_val):
            if not raw_val:
                return None
            apenas_numeros = "".join(c for c in str(raw_val) if c.isdigit())
            if not apenas_numeros:
                return None
            return apenas_numeros.zfill(11)

        def find_col(candidates):
            cols_map = {remove_accents(str(c).strip().upper()): c for c in fieldnames if c is not None}
            for cand in candidates:
                cand_upper = remove_accents(cand.upper())
                if cand_upper in cols_map:
                    return cols_map[cand_upper]
                for real_col_upper, real_col_name in cols_map.items():
                    if cand_upper in real_col_upper:
                        return real_col_name
            return None

        # Força a busca explícita pelo documento e nome do FAMILIAR/BENEFICIÁRIO primeiro
        col_nome = find_col(['Dados do Grupo Familiar > Nome', 'NOME_COMPLETO', 'BENEFICIARIO', 'NOME'])
        col_cpf = find_col(['Dados do Grupo Familiar > CPF', 'CPF_BENEFICIARIO', 'CPF'])
        col_caf = find_col(['DAP / CAF', 'CAF', 'DAP'])
        
        # Isola os dados do técnico para que não colidam
        col_tec = find_col(['Nome do(a) técnico(a) responsável', 'TECNICO', 'RESPONSAVEL'])
        col_mun = find_col(['MUNICIPIO', 'CIDADE'])
        col_com = find_col(['Comunidade', 'LOCALIDADE'])
        
        col_cod_plano = find_col(['Codigo (Plano Produtivo)', 'CODIGO PLANO', 'COD_PLANO', 'CODIGO_PLANO', 'CODIGO'])
        col_objetivo = find_col(['Objetivo', 'OBJETIVO_GERAL', 'OBJETIVO'])
        col_iniciativa = find_col(['Iniciativa', 'INICIATIVA'])
        
        supabase = get_supabase()
        count_novos = 0
        count_atualizados = 0
        linhas_descartadas = [] # Guarda o motivo do descarte de cada linha
        
        # LOG DE DIAGNÓSTICO DE MEMÓRIA BRUTA
        total_linhas_detectadas = len(linhas)
        logger.info(f"🔍 [INVESTIGAÇÃO] O parser leu um total bruto de {total_linhas_detectadas} linhas do arquivo enviado.")
        print(f"🔍 [INVESTIGAÇÃO] QUANTIDADE DE LINHAS ENXERGADAS PELO PYTHON: {total_linhas_detectadas}")
        
        for idx, row in enumerate(linhas, start=2):
            try:
                raw_cpf = row.get(col_cpf) if col_cpf else None
                raw_caf = row.get(col_caf) if col_caf else None
                
                # Passa o dado bruto pela nova inteligência de padronização
                cpf_limpo = padronizar_cpf(raw_cpf)
                
                nome = (row.get(col_nome) or "Desconhecido").strip()
                mun = row.get(col_mun)
                com = row.get(col_com)
                tec = row.get(col_tec)
                
                raw_cod_plano = row.get(col_cod_plano) if col_cod_plano else None
                raw_objetivo = row.get(col_objetivo) if col_objetivo else None
                raw_iniciativa = row.get(col_iniciativa) if col_iniciativa else None
                
                payload = {
                    "nome_completo": nome,
                    "municipio": mun if mun else None,
                    "comunidade": com if com else None,
                    "nome_tecnico": tec if tec else None,
                    "projeto": "Bahia Sem Fome",
                    "cpf": cpf_limpo if cpf_limpo else None,
                    "caf": raw_caf if raw_caf else None
                }
                if raw_cod_plano:
                    payload["codigo_plano"] = raw_cod_plano.strip()
                
                beneficiario_id = None
                existente = None

                # Passo A: Se tiver cpf_limpo, busca por CPF.
                if cpf_limpo:
                    res_check = supabase.table("beneficiarios").select("id, projeto, cpf").eq("cpf", cpf_limpo).execute()
                    if res_check.data:
                        existente = res_check.data[0]
                        beneficiario_id = existente["id"]

                # Passo B: Se não tiver CPF mas tiver raw_caf, busca por CAF.
                if not beneficiario_id and raw_caf:
                    res_check = supabase.table("beneficiarios").select("id, projeto, caf").eq("caf", raw_caf).execute()
                    if res_check.data:
                        existente = res_check.data[0]
                        beneficiario_id = existente["id"]

                # Passo C: Se não tiver nenhum dos dois, busca por correspondência exata de nome_completo e municipio para evitar duplicar.
                if not beneficiario_id:
                    nome_norm = normalizar_nome(nome)
                    mun_norm = normalizar_nome(mun)
                    
                    query_mun = supabase.table("beneficiarios").select("id, nome_completo, municipio, projeto").eq("projeto", "Bahia Sem Fome")
                    if mun:
                        query_mun = query_mun.ilike("municipio", mun.strip())
                    res_mun = query_mun.execute()
                    
                    for b in (res_mun.data or []):
                        b_nome_norm = normalizar_nome(b.get("nome_completo"))
                        b_mun_norm = normalizar_nome(b.get("municipio"))
                        if b_nome_norm == nome_norm and b_mun_norm == mun_norm:
                            existente = b
                            beneficiario_id = b["id"]
                            break

                if beneficiario_id:
                    # Executa a atualização (PATCH) do payload existente
                    try:
                        supabase.table("beneficiarios").update(payload).eq("id", beneficiario_id).execute()
                        count_atualizados += 1
                    except Exception as upd_err:
                        logger.error(f"❌ [ERRO UPDATE] Falha ao atualizar beneficiário ID {beneficiario_id} na linha {idx}: {upd_err}")
                        print(f"❌ [ERRO UPDATE] Falha na linha {idx}: {upd_err}")
                        linhas_descartadas.append({
                            "linha": idx,
                            "nome": nome,
                            "motivo": f"Falha na atualização do registro existente: {str(upd_err)}"
                        })
                else:
                    # SE NÃO FOI ENCONTRADO (Caso das 45 linhas novas), OBRIGATORIAMENTE EXECUTA O INSERT (POST)
                    insert_payload = payload.copy()
                    if cpf_limpo:
                        insert_payload["cpf"] = cpf_limpo
                    insert_payload["status"] = "IMPORTADO"
                    
                    try:
                        res_ins = supabase.table("beneficiarios").insert(insert_payload).execute()
                        if res_ins.data:
                            beneficiario_id = res_ins.data[0]["id"]
                        count_novos += 1
                    except Exception as ins_err:
                        err_msg = str(ins_err).lower()
                        if "duplicate key" in err_msg or "violates unique constraint" in err_msg:
                            logger.warning(f"⚠️ [AVISO INSERT] Beneficiário concorrente ou duplicado na planilha na linha {idx}: {ins_err}")
                            print(f"⚠️ [AVISO INSERT] Duplicidade na linha {idx}: {ins_err}")
                            linhas_descartadas.append({
                                "linha": idx,
                                "nome": nome,
                                "motivo": "Beneficiário com CPF/CAF já existente ou cadastrado concorrentemente."
                            })
                        else:
                            logger.error(f"❌ [ERRO INSERT] Falha ao inserir novo beneficiário na linha {idx}: {ins_err}")
                            print(f"❌ [ERRO INSERT] Falha na linha {idx}: {ins_err}")
                            linhas_descartadas.append({
                                "linha": idx,
                                "nome": nome,
                                "motivo": f"Falha na inserção do registro: {str(ins_err)}"
                            })
                
                if beneficiario_id and raw_cod_plano:
                    cod_iniciativa = raw_cod_plano.strip()
                    partes = cod_iniciativa.split('.')
                    if len(partes) > 1:
                        partes[-1] = '0'
                        cod_objetivo = '.'.join(partes)
                    else:
                        cod_objetivo = cod_iniciativa + ".0"
                        
                    # Processa o OBJETIVO
                    if raw_objetivo:
                        desc_obj = raw_objetivo.strip()
                        res_obj = supabase.table("bsf_metas_plano").select("id, descricao").eq("beneficiario_id", beneficiario_id).eq("codigo", cod_objetivo).eq("tipo", "OBJETIVO").execute()
                        if res_obj.data:
                            obj_id = res_obj.data[0]["id"]
                            if res_obj.data[0]["descricao"] != desc_obj:
                                supabase.table("bsf_metas_plano").update({"descricao": desc_obj}).eq("id", obj_id).execute()
                        else:
                            supabase.table("bsf_metas_plano").insert({
                                "beneficiario_id": beneficiario_id,
                                "codigo": cod_objetivo,
                                "tipo": "OBJETIVO",
                                "descricao": desc_obj
                            }).execute()
                            
                    # Processa a INICIATIVA
                    if raw_iniciativa:
                        desc_ini = raw_iniciativa.strip()
                        res_ini = supabase.table("bsf_metas_plano").select("id, descricao").eq("beneficiario_id", beneficiario_id).eq("codigo", cod_iniciativa).eq("tipo", "INICIATIVA").execute()
                        if res_ini.data:
                            ini_id = res_ini.data[0]["id"]
                            if res_ini.data[0]["descricao"] != desc_ini:
                                supabase.table("bsf_metas_plano").update({"descricao": desc_ini}).eq("id", ini_id).execute()
                        else:
                            supabase.table("bsf_metas_plano").insert({
                                "beneficiario_id": beneficiario_id,
                                "codigo": cod_iniciativa,
                                "tipo": "INICIATIVA",
                                "descricao": desc_ini
                            }).execute()
            except Exception as e:
                error_msg = str(e).lower()
                if 'pgrst204' in error_msg or 'caf' in error_msg:
                    try:
                        supabase.rpc('reload_schema', {}).execute()
                    except Exception:
                        pass
                    return JSONResponse(content={"error": "O esquema do banco de dados foi atualizado. Por favor, tente novamente em 5 segundos."}, status_code=500)
                logger.error(f"❌ [ERRO CRÍTICO LINHA {idx}]: {str(e)}")
                print(f"❌ [ERRO CRÍTICO LINHA {idx}]: {str(e)}")
                linhas_descartadas.append({
                    "linha": idx,
                    "nome": nome if 'nome' in locals() else "Desconhecido",
                    "motivo": f"Erro de runtime: {str(e)}"
                })
                
        return {
            "message": "Importação de beneficiários concluída",
            "novos_inseridos": count_novos,
            "atualizados": count_atualizados,
            "total_linhas_detectadas": total_linhas_detectadas,
            "linhas_ignoradas_count": len(linhas_descartadas),
            "diagnostico_descartes": linhas_descartadas
        }
    except Exception as e:
        logger.error(f"Erro importar bsf: {e}")
        return JSONResponse(content={"error": str(e)}, status_code=500)


@router.post("/importar-metas")
async def importar_planilha_metas_bsf(file: UploadFile = File(...)):
    """Importa metas do PowerBI para beneficiários BSF, sobrescrevendo dados antigos do beneficiário."""
    try:
        content = await file.read()
        filename = file.filename.lower()
        
        linhas = []
        fieldnames = []
        
        if filename.endswith('.xlsx'):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            sheet = wb.active
            
            headers = [cell.value for cell in sheet[1]]
            fieldnames = [str(h).strip() for h in headers if h is not None]
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(row):
                    row_dict = {}
                    for i, header in enumerate(fieldnames):
                        if i < len(row):
                            val = row[i]
                            row_dict[header] = str(val).strip() if val is not None else None
                    linhas.append(row_dict)
                    
        elif filename.endswith('.xls'):
            return JSONResponse(content={"error": "Formato .xls legado não suportado. Use .xlsx ou .csv."}, status_code=400)
            
        else:
            decoded = content.decode('utf-8-sig')
            sniffer = csv.Sniffer()
            try:
                dialect = sniffer.sniff(decoded[:1024], delimiters=[',', ';', '\t'])
                delimiter = dialect.delimiter
            except csv.Error:
                delimiter = ';' if ';' in decoded[:100] else ','
                
            csv_reader = csv.DictReader(io.StringIO(decoded), delimiter=delimiter)
            linhas = list(csv_reader)
            fieldnames = csv_reader.fieldnames or []

        def remove_accents(text):
            if not text:
                return ""
            return "".join(c for c in unicodedata.normalize('NFD', str(text)) if unicodedata.category(c) != 'Mn')

        def find_col(candidates):
            cols_map = {remove_accents(str(c).strip().upper()): c for c in fieldnames if c is not None}
            for cand in candidates:
                cand_upper = remove_accents(cand.upper())
                if cand_upper in cols_map:
                    return cols_map[cand_upper]
                for real_col_upper, real_col_name in cols_map.items():
                    if cand_upper in real_col_upper:
                        return real_col_name
            return None

        col_codigo = find_col(['Codigo (Plano Produtivo)', 'CODIGO PLANO', 'COD_PLANO', 'CODIGO_PLANO', 'CODIGO', 'CHAVE'])
        
        if not col_codigo:
            return JSONResponse(content={"error": "Coluna de código de plano/metas não encontrada na planilha."}, status_code=400)

        # Colunas adicionais usando find_col de forma resiliente
        col_cpf = find_col(['CPF', 'DOCUMENTO', 'DOC', 'CPF_BENEFICIARIO'])
        col_nome = find_col(['NOME', 'BENEFICIARIO', 'NOME COMPLETO', 'NOME_BENEFICIARIO'])
        col_tarefa = find_col(['TAREFA', 'ATIVIDADE_TAREFA', 'TAREFA_DESCRICAO'])
        col_como_fazer = find_col(['COMO FAZER', 'COMO_FAZER', 'METODO', 'ACAO'])

        supabase = get_supabase()
        count_metas_inseridas = 0
        beneficiarios_limpos = set()
        warnings = []
        
        for line_num, row in enumerate(linhas, start=2):
            raw_codigo_str = "N/A"
            try:
                raw_codigo = row.get(col_codigo)
                if not raw_codigo:
                    continue
                
                raw_codigo_str = str(raw_codigo).strip()
                partes = raw_codigo_str.split('.')
                if not partes or not partes[0]:
                    continue
                
                codigo_plano_extraido = partes[0].strip()
                
                # Classifica o Tipo (um ponto = OBJETIVO; dois pontos = INICIATIVA)
                pontos = raw_codigo_str.count('.')
                if pontos == 1:
                    tipo = "OBJETIVO"
                elif pontos >= 2:
                    tipo = "INICIATIVA"
                else:
                    continue
                
                # Extração resiliente de dados do beneficiário
                raw_cpf = row.get(col_cpf) if col_cpf else None
                cpf_limpo = limpar_cpf(raw_cpf) if raw_cpf else None
                raw_nome = row.get(col_nome) if col_nome else None
                nome_limpo = raw_nome.strip() if raw_nome else None
                
                beneficiario_id = None
                
                # 1. Match por CPF primeiro (se fornecido)
                if cpf_limpo:
                    res_ben = supabase.table("beneficiarios").select("id").eq("cpf", cpf_limpo).eq("projeto", "Bahia Sem Fome").execute()
                    if res_ben.data:
                        beneficiario_id = res_ben.data[0]["id"]
                
                # 2. Match por Nome Completo segundo (case-insensitive)
                if not beneficiario_id and nome_limpo:
                    res_ben = supabase.table("beneficiarios").select("id").ilike("nome_completo", nome_limpo).eq("projeto", "Bahia Sem Fome").execute()
                    if res_ben.data:
                        beneficiario_id = res_ben.data[0]["id"]
                
                # 3. Match por codigo_plano atual
                if not beneficiario_id and codigo_plano_extraido:
                    res_ben = supabase.table("beneficiarios").select("id").eq("codigo_plano", codigo_plano_extraido).eq("projeto", "Bahia Sem Fome").execute()
                    if res_ben.data:
                        beneficiario_id = res_ben.data[0]["id"]
                
                # Se não encontrado por nenhuma das formas:
                if not beneficiario_id:
                    warnings.append({
                        "linha": line_num,
                        "codigo": raw_codigo_str,
                        "erro": f"Beneficiário não encontrado para CPF: {raw_cpf or 'N/A'}, Nome: {raw_nome or 'N/A'}, Código do Plano: {codigo_plano_extraido}."
                    })
                    continue
                
                # Atualizar o codigo_plano na tabela beneficiarios com a parte puramente numérica
                if codigo_plano_extraido:
                    supabase.table("beneficiarios").update({"codigo_plano": codigo_plano_extraido}).eq("id", beneficiario_id).execute()
                
                # Sobrescrita exclusiva: limpa as metas na primeira aparição do beneficiário
                if beneficiario_id not in beneficiarios_limpos:
                    supabase.table("bsf_metas_plano").delete().eq("beneficiario_id", beneficiario_id).execute()
                    beneficiarios_limpos.add(beneficiario_id)
                
                # Procura descrição de forma flexível de acordo com o tipo
                raw_descricao = None
                colunas_descricao = ['OBJETIVO', 'INICIATIVA', 'DESCRICAO', 'TEXTO', 'META']
                for cand in colunas_descricao:
                    col_cand = find_col([cand])
                    if col_cand and row.get(col_cand):
                        raw_descricao = str(row.get(col_cand)).strip()
                        break
                
                raw_tarefa = str(row.get(col_tarefa)).strip() if col_tarefa and row.get(col_tarefa) else None
                raw_como_fazer = str(row.get(col_como_fazer)).strip() if col_como_fazer and row.get(col_como_fazer) else None
                
                # Insere a nova meta limpa
                payload_meta = {
                    "beneficiario_id": beneficiario_id,
                    "codigo": raw_codigo_str,
                    "tipo": tipo,
                    "descricao": raw_descricao or "",
                    "tarefa": raw_tarefa or "",
                    "como_fazer": raw_como_fazer or ""
                }
                supabase.table("bsf_metas_plano").insert(payload_meta).execute()
                count_metas_inseridas += 1
                
            except Exception as e:
                logger.error(f"Erro ao processar linha de metas {line_num}: {e}")
                warnings.append({
                    "linha": line_num,
                    "codigo": raw_codigo_str,
                    "erro": f"Erro inesperado: {str(e)}"
                })
                
        return {
            "message": "Importação de metas concluída com sucesso",
            "warnings": warnings,
            "metas_inseridas": count_metas_inseridas,
            "beneficiarios_atualizados": len(beneficiarios_limpos)
        }
    except Exception as e:
        logger.error(f"Erro ao importar metas BSF: {e}")
        return JSONResponse(content={"error": str(e)}, status_code=500)


# --- NOVOS ENDPOINTS DO WIZARD DE IMPORTAÇÃO BSF (PASSOS 2, 3 E 4) ---

async def _ler_planilha_upload(file: UploadFile) -> tuple[list, list]:
    content = await file.read()
    filename = file.filename.lower()
    linhas = []
    fieldnames = []
    
    if filename.endswith('.xlsx'):
        import openpyxl
        from zipfile import BadZipFile
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            sheet = wb.active
        except BadZipFile as bz:
            logger.error(f"Arquivo XLSX corrompido fisicamente: {bz}")
            raise HTTPException(status_code=400, detail=f"Arquivo Excel corrompido (Zip inválido): {str(bz)}")
        except Exception as ex:
            logger.error(f"Erro ao ler Excel: {ex}")
            raise HTTPException(status_code=400, detail=f"Erro ao ler Excel: {str(ex)}")
        
        headers = [cell.value for cell in sheet[1]]
        fieldnames = [str(h).strip() for h in headers if h is not None]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(row):
                row_dict = {}
                for i, header in enumerate(fieldnames):
                    if i < len(row):
                        val = row[i]
                        row_dict[header] = str(val).strip() if val is not None else None
                linhas.append(row_dict)
                
    elif filename.endswith('.xls'):
        raise HTTPException(status_code=400, detail="Formato .xls legado não suportado. Use .xlsx ou .csv.")
        
    else:
        try:
            decoded = content.decode('utf-8-sig')
        except UnicodeDecodeError:
            try:
                decoded = content.decode('iso-8859-1')
            except Exception:
                raise HTTPException(status_code=400, detail="Erro de codificação de caracteres no CSV.")
                
        try:
            sniffer = csv.Sniffer()
            try:
                dialect = sniffer.sniff(decoded[:1024], delimiters=[',', ';', '\t'])
                delimiter = dialect.delimiter
            except Exception:
                delimiter = ';' if ';' in decoded[:100] else ','
            
            csv_reader = csv.DictReader(io.StringIO(decoded), delimiter=delimiter)
            linhas = list(csv_reader)
            fieldnames = csv_reader.fieldnames or []
        except Exception as csv_err:
            raise HTTPException(status_code=400, detail=f"Erro na estrutura interna do arquivo CSV: {str(csv_err)}")
            
    return linhas, fieldnames

def remove_accents(text):
    if not text:
        return ""
    return "".join(c for c in unicodedata.normalize('NFD', str(text)) if unicodedata.category(c) != 'Mn')

def find_col_in_list(fieldnames, candidates):
    cols_map = {remove_accents(str(c).strip().upper()): c for c in fieldnames if c is not None}
    for cand in candidates:
        cand_upper = remove_accents(cand.upper())
        if cand_upper in cols_map:
            return cols_map[cand_upper]
        for real_col_upper, real_col_name in cols_map.items():
            if cand_upper in real_col_upper:
                return real_col_name
    return None

@router.post("/vincular-codigos")
async def vincular_codigos_plano(file: UploadFile = File(...)):
    """Passo 2: Vincula o código base do plano produtivo ao beneficiário por CPF ou Nome."""
    try:
        linhas, fieldnames = await _ler_planilha_upload(file)
        
        col_cpf = find_col_in_list(fieldnames, ['CPF', 'DOCUMENTO', 'DOC', 'CPF_BENEFICIARIO'])
        col_nome = find_col_in_list(fieldnames, ['Dados do Grupo Familiar > Nome', 'NOME_COMPLETO', 'BENEFICIARIO', 'NOME'])
        col_cod_plano = find_col_in_list(fieldnames, ['Codigo (Plano Produtivo)', 'CODIGO PLANO', 'COD_PLANO', 'CODIGO_PLANO', 'CODIGO', 'CHAVE'])
        
        if not col_cod_plano:
            return JSONResponse(content={"error": "Coluna de código do plano não encontrada na planilha."}, status_code=400)
            
        supabase = get_supabase()
        count_sucesso = 0
        warnings = []
        
        def padronizar_cpf(raw_val):
            if not raw_val:
                return None
            apenas_numeros = "".join(c for c in str(raw_val) if c.isdigit())
            if not apenas_numeros:
                return None
            return apenas_numeros.zfill(11)
            
        for line_num, row in enumerate(linhas, start=2):
            raw_codigo = row.get(col_cod_plano)
            if not raw_codigo:
                continue
                
            raw_codigo_str = str(raw_codigo).strip()
            # Extrair a parte base numérica do código do plano (ex: 31646.1.1 -> 31646)
            codigo_plano_extraido = raw_codigo_str.split('.')[0].strip()
            if not codigo_plano_extraido:
                continue
                
            raw_cpf = row.get(col_cpf) if col_cpf else None
            cpf_limpo = padronizar_cpf(raw_cpf) if raw_cpf else None
            raw_nome = row.get(col_nome) if col_nome else None
            nome_limpo = raw_nome.strip() if raw_nome else None
            
            beneficiario_id = None
            
            try:
                # 1. Busca por CPF
                if cpf_limpo:
                    res_ben = supabase.table("beneficiarios").select("id").eq("cpf", cpf_limpo).eq("projeto", "Bahia Sem Fome").execute()
                    if res_ben.data:
                        beneficiario_id = res_ben.data[0]["id"]
                        
                # 2. Busca por Nome
                if not beneficiario_id and nome_limpo:
                    res_ben = supabase.table("beneficiarios").select("id").ilike("nome_completo", nome_limpo).eq("projeto", "Bahia Sem Fome").execute()
                    if res_ben.data:
                        beneficiario_id = res_ben.data[0]["id"]
                        
                if not beneficiario_id:
                    warnings.append({
                        "linha": line_num,
                        "codigo": raw_codigo_str,
                        "erro": f"Beneficiário não encontrado para CPF: {raw_cpf or 'N/A'}, Nome: {raw_nome or 'N/A'}."
                    })
                    continue
                    
                # Atualizar código do plano no beneficiário correspondente
                supabase.table("beneficiarios").update({"codigo_plano": codigo_plano_extraido}).eq("id", beneficiario_id).execute()
                count_sucesso += 1
                
            except Exception as e:
                logger.error(f"Erro ao processar Passo 2 na linha {line_num}: {e}")
                warnings.append({
                    "linha": line_num,
                    "codigo": raw_codigo_str,
                    "erro": f"Erro inesperado: {str(e)}"
                })
                
        return {
            "message": "Vinculação de códigos concluída",
            "sucesso": count_sucesso,
            "warnings": warnings
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Erro ao importar Passo 2: {e}")
        return JSONResponse(content={"error": str(e)}, status_code=500)

@router.post("/importar-objetivos")
async def importar_objetivos(file: UploadFile = File(...)):
    """Passo 3: Importa objetivos e metas salvando-os como nós pais na tabela de metas."""
    try:
        linhas, fieldnames = await _ler_planilha_upload(file)
        
        col_codigo = find_col_in_list(fieldnames, ['Codigo (Plano Produtivo)', 'CODIGO PLANO', 'COD_PLANO', 'CODIGO_PLANO', 'CODIGO', 'CHAVE'])
        col_objetivo = find_col_in_list(fieldnames, ['Objetivo', 'OBJETIVO_GERAL', 'OBJETIVO'])
        col_meta = find_col_in_list(fieldnames, ['Meta', 'META_DESCRICAO', 'INICIATIVA', 'DESCRICAO'])
        
        if not col_codigo:
            return JSONResponse(content={"error": "Coluna de código do plano não encontrada na planilha."}, status_code=400)
            
        supabase = get_supabase()
        count_sucesso = 0
        beneficiarios_limpos = set()
        warnings = []
        
        for line_num, row in enumerate(linhas, start=2):
            raw_codigo = row.get(col_codigo)
            if not raw_codigo:
                continue
                
            raw_codigo_str = str(raw_codigo).strip()
            partes = raw_codigo_str.split('.')
            if not partes or not partes[0]:
                continue
                
            codigo_plano_extraido = partes[0].strip()
            
            # Buscar beneficiário pelo código do plano
            res_ben = supabase.table("beneficiarios").select("id").eq("codigo_plano", codigo_plano_extraido).eq("projeto", "Bahia Sem Fome").execute()
            if not res_ben.data:
                warnings.append({
                    "linha": line_num,
                    "codigo": raw_codigo_str,
                    "erro": f"Beneficiário não cadastrado para o código de plano: {codigo_plano_extraido}."
                })
                continue
                
            beneficiario_ids = [b["id"] for b in res_ben.data]
            
            raw_objetivo = row.get(col_objetivo)
            raw_meta = row.get(col_meta)
            
            desc_obj = str(raw_objetivo).strip() if raw_objetivo else None
            desc_meta = str(raw_meta).strip() if raw_meta else None
            
            # Formatação do código do objetivo correspondente (ex: 31646.1.1 -> 31646.1.0)
            if len(partes) > 1:
                partes_obj = partes.copy()
                partes_obj[-1] = '0'
                cod_objetivo = '.'.join(partes_obj)
            else:
                cod_objetivo = raw_codigo_str + ".0"
                
            cod_iniciativa = raw_codigo_str
            
            try:
                for ben_id in beneficiario_ids:
                    # Limpeza das metas na primeira aparição do beneficiário nesta importação (sobrescrita limpa)
                    if ben_id not in beneficiarios_limpos:
                        supabase.table("bsf_metas_plano").delete().eq("beneficiario_id", ben_id).execute()
                        beneficiarios_limpos.add(ben_id)
                        
                    # 1. Inserir Objetivo (nó pai - tipo OBJETIVO)
                    if desc_obj:
                        res_obj_check = supabase.table("bsf_metas_plano").select("id").eq("beneficiario_id", ben_id).eq("codigo", cod_objetivo).eq("tipo", "OBJETIVO").execute()
                        if not res_obj_check.data:
                            supabase.table("bsf_metas_plano").insert({
                                "beneficiario_id": ben_id,
                                "codigo": cod_objetivo,
                                "tipo": "OBJETIVO",
                                "descricao": desc_obj
                            }).execute()
                            
                    # 2. Inserir Meta/Iniciativa (nó pai - tipo INICIATIVA)
                    if desc_meta:
                        res_meta_check = supabase.table("bsf_metas_plano").select("id").eq("beneficiario_id", ben_id).eq("codigo", cod_iniciativa).eq("tipo", "INICIATIVA").execute()
                        if not res_meta_check.data:
                            supabase.table("bsf_metas_plano").insert({
                                "beneficiario_id": ben_id,
                                "codigo": cod_iniciativa,
                                "tipo": "INICIATIVA",
                                "descricao": desc_meta
                            }).execute()
                            
                count_sucesso += 1
                
            except Exception as e:
                logger.error(f"Erro ao processar Passo 3 na linha {line_num}: {e}")
                warnings.append({
                    "linha": line_num,
                    "codigo": raw_codigo_str,
                    "erro": f"Erro inesperado: {str(e)}"
                })
                
        return {
            "message": "Importação de objetivos concluída com sucesso",
            "sucesso": count_sucesso,
            "warnings": warnings
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Erro ao importar Passo 3: {e}")
        return JSONResponse(content={"error": str(e)}, status_code=500)

@router.post("/importar-iniciativas")
async def importar_iniciativas(file: UploadFile = File(...)):
    """Passo 4: Processa colunas de execução e acopla os dados na tabela por chave composta (codigo + beneficiario_id)."""
    try:
        linhas, fieldnames = await _ler_planilha_upload(file)
        
        col_codigo = find_col_in_list(fieldnames, ['Codigo (Plano Produtivo)', 'CODIGO PLANO', 'COD_PLANO', 'CODIGO_PLANO', 'CODIGO', 'CHAVE'])
        col_tarefa = find_col_in_list(fieldnames, ['TAREFA', 'ATIVIDADE_TAREFA', 'TAREFA_DESCRICAO'])
        col_como_fazer = find_col_in_list(fieldnames, ['COMO FAZER', 'COMO_FAZER', 'METODO', 'ACAO'])
        col_entrega = find_col_in_list(fieldnames, ['ENTREGA', 'PRODUTO', 'ENTREGA_DESCRICAO'])
        col_recursos = find_col_in_list(fieldnames, ['RECURSOS NECESSARIOS', 'RECURSOS', 'RECURSOS_NECESSARIOS'])
        
        if not col_codigo:
            return JSONResponse(content={"error": "Coluna de código não encontrada na planilha."}, status_code=400)
            
        supabase = get_supabase()
        count_sucesso = 0
        warnings = []
        
        for line_num, row in enumerate(linhas, start=2):
            raw_codigo = row.get(col_codigo)
            if not raw_codigo:
                continue
                
            raw_codigo_str = str(raw_codigo).strip()
            partes = raw_codigo_str.split('.')
            if not partes or not partes[0]:
                continue
                
            codigo_plano_extraido = partes[0].strip()
            
            # Buscar beneficiário pelo código do plano
            res_ben = supabase.table("beneficiarios").select("id").eq("codigo_plano", codigo_plano_extraido).eq("projeto", "Bahia Sem Fome").execute()
            if not res_ben.data:
                warnings.append({
                    "linha": line_num,
                    "codigo": raw_codigo_str,
                    "erro": f"Beneficiário com código de plano {codigo_plano_extraido} não encontrado."
                })
                continue
                
            beneficiario_ids = [b["id"] for b in res_ben.data]
            
            raw_tarefa = row.get(col_tarefa)
            raw_como_fazer = row.get(col_como_fazer)
            raw_entrega = row.get(col_entrega)
            raw_recursos = row.get(col_recursos)
            
            desc_tarefa = str(raw_tarefa).strip() if raw_tarefa else None
            desc_como_fazer = str(raw_como_fazer).strip() if raw_como_fazer else None
            desc_entrega = str(raw_entrega).strip() if raw_entrega else None
            desc_recursos = str(raw_recursos).strip() if raw_recursos else None
            
            try:
                for ben_id in beneficiario_ids:
                    # Validar estritamente se a meta pai (tipo INICIATIVA) já existe para o beneficiário (Opção A)
                    res_meta = supabase.table("bsf_metas_plano").select("id").eq("beneficiario_id", ben_id).eq("codigo", raw_codigo_str).eq("tipo", "INICIATIVA").execute()
                    
                    if not res_meta.data:
                        warnings.append({
                            "linha": line_num,
                            "codigo": raw_codigo_str,
                            "erro": f"Meta/Iniciativa pai não encontrada para o beneficiário ID {ben_id}. Passo 3 foi pulado?"
                        })
                        continue
                        
                    meta_id = res_meta.data[0]["id"]
                    
                    # Atualizar com os dados adicionais de execução (acoplamento)
                    payload_update = {}
                    if desc_tarefa is not None:
                        payload_update["tarefa"] = desc_tarefa
                    if desc_como_fazer is not None:
                        payload_update["como_fazer"] = desc_como_fazer
                    if desc_entrega is not None:
                        payload_update["entrega"] = desc_entrega
                    if desc_recursos is not None:
                        payload_update["recursos_necessarios"] = desc_recursos
                        
                    if payload_update:
                        supabase.table("bsf_metas_plano").update(payload_update).eq("id", meta_id).execute()
                        count_sucesso += 1
                        
            except Exception as e:
                logger.error(f"Erro ao processar Passo 4 na linha {line_num}: {e}")
                warnings.append({
                    "linha": line_num,
                    "codigo": raw_codigo_str,
                    "erro": f"Erro inesperado: {str(e)}"
                })
                
        return {
            "message": "Acoplamento de iniciativas concluído",
            "sucesso": count_sucesso,
            "warnings": warnings
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"Erro ao importar Passo 4: {e}")
        return JSONResponse(content={"error": str(e)}, status_code=500)

