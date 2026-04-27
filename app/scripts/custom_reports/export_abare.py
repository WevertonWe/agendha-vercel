import sqlite3
import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paths
BASE_DIR = Path(__file__).resolve().parent.parent.parent.parent
DB_PATH = BASE_DIR / "agendha.db"
OUTPUT_DIR = BASE_DIR / "app" / "static" / "reports"

REPORT_GERAL = OUTPUT_DIR / "abare_geral.xlsx"
REPORT_GRH = OUTPUT_DIR / "abare_grh.xlsx"

def apply_formatting(writer, df, sheet_name="Sheet1"):
    """
    Applies professional formatting using openpyxl:
    - Dark Blue Header with White Text
    - Auto-filter
    - Auto-adjust column width
    """
    workbook = writer.book  # noqa: F841
    worksheet = writer.sheets[sheet_name]
    
    # Define Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Side(border_style="thin", color="000000")
    border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    # Apply Header Styles
    for col_num, value in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
                
    # Auto-adjust column widths
    for i, col in enumerate(df.columns):
        # Calculate max length
        max_len = max(
            df[col].astype(str).map(len).max() if not df[col].empty else 0,
            len(str(col))
        )
        col_letter = get_column_letter(i + 1)
        worksheet.column_dimensions[col_letter].width = max_len + 5
        
    # Add Auto-filter
    worksheet.auto_filter.ref = worksheet.dimensions

def generate_reports():
    conn = sqlite3.connect(DB_PATH)
    
    # Common Columns mapped to Human-Readable Headers
    cols_map = {
        "nome_familiar": "Nome Familiar",
        "cpf_familiar": "CPF",
        "latitude": "Latitude",
        "longitude": "Longitude",
        "grh": "GRH",
        "verificado_bsf": "Verificação BSF"
    }
    
    columns_sql = ", ".join(cols_map.keys())
    
    try:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        
        # --- REPORT A: GERAL (Sem Filtro GRH) ---
        query_geral = f"SELECT {columns_sql} FROM beneficiarios WHERE municipio LIKE '%ABAR%'"  # nosec
        df_geral = pd.read_sql_query(query_geral, conn)
        df_geral.rename(columns=cols_map, inplace=True)
        
        with pd.ExcelWriter(REPORT_GERAL, engine='openpyxl') as writer:
            df_geral.to_excel(writer, index=False, sheet_name='Geral')
            apply_formatting(writer, df_geral, sheet_name='Geral')
            
        print(f"✅ Relatório Geral: {len(df_geral)} registros -> {REPORT_GERAL}")

        # --- REPORT B: APENAS GRH ---
        query_grh = f"""
            SELECT {columns_sql} FROM beneficiarios 
            WHERE municipio LIKE '%ABAR%' 
            AND (grh IS NOT NULL AND grh != '')
        """
        df_grh = pd.read_sql_query(query_grh, conn)
        df_grh.rename(columns=cols_map, inplace=True)
        
        # Even if empty, generate file with headers
        with pd.ExcelWriter(REPORT_GRH, engine='openpyxl') as writer:
            df_grh.to_excel(writer, index=False, sheet_name='GRH_Preenchido')
            apply_formatting(writer, df_grh, sheet_name='GRH_Preenchido')

        print(f"✅ Relatório GRH: {len(df_grh)} registros -> {REPORT_GRH}")
        
    except Exception as e:
        print(f"❌ Erro ao gerar relatórios: {e}")
    finally:
        conn.close()

if __name__ == "__main__":
    generate_reports()
